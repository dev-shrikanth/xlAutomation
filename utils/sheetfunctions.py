import openpyxl


def get_sheet_and_range(wb, named_range):
    ws, range_address = wb.defined_names[named_range].attr_text.split("!")
    return ws, range_address


def vlookup(lookup_value, return_column_index, ws, lookup_range):

    if type(lookup_value) == str:
        for row in ws[lookup_range]:
            if str(row[0].value) == lookup_value:
                return row[return_column_index - 1].value
        return None
    else:
        for row in ws[lookup_range]:
            if row[0].value == lookup_value:
                return row[return_column_index - 1].value
        return None


def vlookup_number(lookup_value, return_column_index, ws, lookup_range):
    for row in ws[lookup_range]:
        if row[0].value == lookup_value:
            return row[return_column_index - 1].value
    return None


# function to traverse through the rows of a lookup_range and return the value from the return_column_index in an array
def vlookup_populate_list(return_column_index, ws, lookup_range):
    retval = []
    for row in ws[lookup_range]:
        retval.append(str(row[return_column_index - 1].value))
    return retval


def set_formulae(wb: openpyxl.Workbook):
    ws, rg = get_sheet_and_range(wb, "f_revenue")
    ws_formula_src = wb[ws]

    ws_revenue = wb["Revenue"]
    for row in ws_formula_src[rg]:
        if row[0].row != 1:
            # read cell address from the first column and set the formula in the cell address from the second column
            cell = row[0].value
            val = row[1].value
            ws_revenue[row[0].value] = row[1].value
