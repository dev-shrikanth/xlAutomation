import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Commiting to branch 2 


def vlookup(lookup_value,  return_column_index, ws, lookup_range): 

    if type(lookup_value) == str:
      for row in ws[lookup_range]:
          if str(row[0].value)  == lookup_value:
              return row[return_column_index - 1].value
      return None
    else:
      for row in ws[lookup_range]:
          if row[0].value  == lookup_value:
              return row[return_column_index - 1].value
      return None

def vlookup_number(lookup_value,  return_column_index, ws, lookup_range): 
    for row in ws[lookup_range]:
        if row[0].value  == lookup_value:
            return row[return_column_index - 1].value
    return None
# function to traverse through the rows of a lookup_range and return the value from the return_column_index in an array
def vlookup_populate_list(return_column_index, ws, lookup_range):
    retval = []
    for row in ws[lookup_range]:
        retval.append(str(row[return_column_index - 1].value))
    return retval

wb = openpyxl.load_workbook('output.xlsx')

# get the named range called 'albums' in this workbook
ws_source = wb['backend']
named_range = wb.defined_names['albums']
ws, range_address = named_range.attr_text.split('!')


# create a worksheet called 'output'
ws_dest = wb['Revenue']

# Loop through the rows of the range_address and populate the ws_dest with the values from the first column of the range_address 
for row in ws_source[range_address]:
    # ignore the first title row
    if row[0].row != 1:
        # ws_dest.append([row[0].value])
        #set value in B column from second row onwards  
        ws_dest['B' + str(row[0].row)] = row[0].value
        ws_dest['C' + str(row[0].row)] = row[1].value
        ws_dest['E' + str(row[0].row)] = row[2].value
        


# get the named range called 'albums' in this workbook
ws_source = wb['backend']
named_range = wb.defined_names['albums']
ws, range_address = named_range.attr_text.split('!')

def set_formulae():
    ws_formula_src = wb['Formulae']
    named_rg = wb.defined_names['f_revenue']
    ws, rg = named_rg.attr_text.split('!')
    ws_revenue = wb['Revenue']
    for row in ws_formula_src[rg]:
        if row[0].row != 1:
            # read cell address from the first column and set the formula in the cell address from the second column
            cell = row[0].value
            val = row[1].value
            ws_revenue[row[0].value] = row[1].value

            


# ws_dest['A2'] = 10
# ws_dest['C2'] = '=VLOOKUP(A2, albums, 2, FALSE)'

# # set data validation as list for the cell ws_dest['A2'] with the values from vlookup_populate_list(1, ws_source, range_address)
# lov = vlookup_populate_list(1, ws_source, range_address)
# lov_string = '"' + ','.join(lov) + '"'
# dv = DataValidation(type="list", formula1=lov_string)

# # add the data validation to the cell ws_dest['A2']
# ws_dest.add_data_validation(dv)
# dv.add(ws_dest['A3'])

# perform a vlookup on the range_address with value from ws_dest['A2'] and return the value from the 2nd column from the range_address and put it in ws_dest['B2']
# retval = vlookup(ws_dest['A2'].value, 3, ws_source, range_address)
# ws_dest['B2'] = retval
set_formulae()

wb.save('output1.xlsx')

# Committed on 2021-01-01 17:00:00



